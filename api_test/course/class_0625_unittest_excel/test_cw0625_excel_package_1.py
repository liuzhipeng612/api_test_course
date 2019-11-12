#!/usr/bin/env python
# coding=UTF-8
"""
@Author: Jomer
@Contact: jomer3126@gmail.com
@Project: Jomer
@File: test_cw0625_excel_package_1.py
@Time: 2019-06-30 11:44
@Desc: Jungle old monster
"""
from openpyxl import load_workbook  # 导入openpyxl第三方库


class HandleExcel:  # 创建一个Excel处理类
    def __init__(self, filename, sheetname=None):  # 初始化类，并设置两个参数
        self.filename = filename
        self.sheetname = sheetname

    def get_case(self):  # 获取行-测试用例
        wb = load_workbook(self.filename)  # 加载Excel文件
        if self.sheetname is None:  # 判断是否有指定表单，如果没有使用打开Excel文件的激活表单，如果有就是用指定名称的表单
            ws = wb.active
        else:
            ws = wb[self.sheetname]

        cases_list = []  # 创建一个空的用例列表
        for row in range(2, ws.max_row + 1):  # 获取Excel的每一行
            case_id = ws.cell(row, 1).value  # 获取Excel当前行的对应单元格的值
            title = ws.cell(row, 2).value
            l_data = ws.cell(row, 3).value
            r_data = ws.cell(row, 4).value
            expected = ws.cell(row, 5).value
            cases_list.append({  # 将获取Excel当前行对应的单元格的值使用一个字典的形式添加在空用例列表中
                'case_id': case_id,
                'title': title,
                'l_data': l_data,
                'r_data': r_data,
                'expected': expected
            })
        return cases_list  # 将用例列表返回给当前函数


if __name__ == '__main__':  # 在当前路径下才可以运行
    file_name = 'demo.xlsx'  # 指定文件
    my_excel = HandleExcel(file_name)  # 实例化my_excel这个对象
    print(my_excel.get_case())  # 将对象调用方法后的结果打印出来
