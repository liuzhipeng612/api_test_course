#!/usr/bin/env python
# coding=UTF-8
"""
@Author: Jomer
@Contact: jomer3126@gmail.com
@Project: Jomer
@File: test_cw0625_case_1.py
@Time: 2019-06-29 22:18
@Desc: Jungle old monster
"""
import unittest

from openpyxl import load_workbook

from interface_automation.class_0625_unittest_excel.cw0625_testing_object import Arithmetic
from interface_automation.class_0625_unittest_excel.test_cw0625_excel_package_2 import HandleExcel


class TestArithmetic(unittest.TestCase):
    file_name_1 = 'demo.txt'
    file_name_2 = 'demo.xlsx'

    @classmethod
    def setUpClass(cls):
        print('\n{:=^40s}\n'.format('开始执行用例'))
        print('打开【{}】文件\n'.format(cls.file_name_1))
        cls.file_1 = open(cls.file_name_1, 'a', encoding='utf-8')
        cls.file_2 = load_workbook(cls.file_name_2)
        cls.file_1.write('\n{:=^40s}\n'.format('开始执行用例'))

    @classmethod
    def tearDownClass(cls):
        print('关闭【{}】文件'.format(cls.file_name_1))
        print('\n{:=^40s}\n'.format('用例执行结束'))
        cls.file_1.write('\n{:=^40s}\n'.format('用例执行结束'))
        cls.file_1.close()
        cls.file_2.save(cls.file_name_2)
        cls.file_2.close()

    def test_add(self):
        self.ws_2 = self.file_2['add']
        my_excel = HandleExcel(self.file_name_2, 'add')
        for index, case_list in enumerate(my_excel.get_case()):
            real_result = Arithmetic(case_list['l_data'], case_list['r_data']).add()
            expect_result = case_list['expected']
            msg = case_list['title']
            try:
                self.assertEqual(expect_result, real_result, msg)
                print('\n{},执行结果为:{}\n'.format(msg, 'Pass'))
                self.file_1.write('\n{},执行结果为:{}\n'.format(msg, 'Pass'))
                self.ws_2.cell(index + 2, 6).value = real_result
                self.ws_2.cell(index + 2, 7).value = 'Pass'
            except AssertionError as e:
                print('具体异常为{}'.format(e))
                self.file_1.write('\n{},执行结果为:{}\n具体异常为:{}\n'.format(msg, 'Fail', e))
                self.ws_2.cell(index + 2, 6).value = real_result
                self.ws_2.cell(index + 2, 7).value = 'Fail'
                raise e

    def test_minus(self):
        self.ws_2 = self.file_2['minus']
        my_excel = HandleExcel(self.file_name_2, 'minus')
        for index, case_list in enumerate(my_excel.get_case()):
            real_result = Arithmetic(case_list['l_data'], case_list['r_data']).minus()
            expect_result = case_list['expected']
            msg = case_list['title']
            try:
                self.assertEqual(expect_result, real_result, msg)
                print('\n{},执行结果为:{}\n'.format(msg, 'Pass'))
                self.file_1.write('\n{},执行结果为:{}\n'.format(msg, 'Pass'))
                self.ws_2.cell(index + 2, 6).value = real_result
                self.ws_2.cell(index + 2, 7).value = 'Pass'
            except AssertionError as e:
                print('具体异常为{}'.format(e))
                self.file_1.write('\n{},执行结果为:{}\n具体异常为:{}\n'.format(msg, 'Fail', e))
                self.ws_2.cell(index + 2, 6).value = real_result
                self.ws_2.cell(index + 2, 7).value = 'Fail'
                raise e

    def test_multiply(self):
        self.ws_2 = self.file_2['multiply']
        my_excel = HandleExcel(self.file_name_2, 'multiply')
        for index, case_list in enumerate(my_excel.get_case()):
            real_result = Arithmetic(case_list['l_data'], case_list['r_data']).multiply()
            expect_result = case_list['expected']
            msg = case_list['title']
            try:
                self.assertEqual(expect_result, real_result, msg)
                print('\n{},执行结果为:{}\n'.format(msg, 'Pass'))
                self.file_1.write('\n{},执行结果为:{}\n'.format(msg, 'Pass'))
                self.ws_2.cell(index + 2, 6).value = real_result
                self.ws_2.cell(index + 2, 7).value = 'Pass'
            except AssertionError as e:
                print('具体异常为{}'.format(e))
                self.file_1.write('\n{},执行结果为:{}\n具体异常为:{}\n'.format(msg, 'Fail', e))
                self.ws_2.cell(index + 2, 6).value = real_result
                self.ws_2.cell(index + 2, 7).value = 'Fail'
                raise e

    def test_division(self):
        self.ws_2 = self.file_2['division']
        my_excel = HandleExcel(self.file_name_2, 'division')
        for index, case_list in enumerate(my_excel.get_case()):
            real_result = Arithmetic(case_list['l_data'], case_list['r_data']).division()
            expect_result = case_list['expected']
            msg = case_list['title']
            try:
                self.assertEqual(expect_result, real_result, msg)
                print('\n{},执行结果为:{}\n'.format(msg, 'Pass'))
                self.file_1.write('\n{},执行结果为:{}\n'.format(msg, 'Pass'))
                self.ws_2.cell(index + 2, 6).value = real_result
                self.ws_2.cell(index + 2, 7).value = 'Pass'
            except AssertionError as e:
                print('具体异常为{}'.format(e))
                self.file_1.write('\n{},执行结果为:{}\n具体异常为:{}\n'.format(msg, 'Fail', e))
                self.ws_2.cell(index + 2, 6).value = real_result
                self.ws_2.cell(index + 2, 7).value = 'Fail'
                raise e


if __name__ == '__main__':
    unittest.main()
