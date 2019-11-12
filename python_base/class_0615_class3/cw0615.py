#!/usr/bin/env python
# coding=UTF-8
"""
@Author: Cautious Eureka
@Contact: Eureka@Faisen.com
@Project: cautious_eureka
@File: cw0615.py
@Time: 2019-06-23 21:57
@Desc: Python full stack automation test engineer(Cautious Eureka)
"""

from openpyxl import load_workbook, Workbook


# 1、总结类和对象的所有知识点。思维导图形式或者文字。

# 2、在之前定义的手机类下面定义智能手机类和苹果手机类。智能手机打电话具有录音功能。 苹果手机打电话不仅具有录音功能，还有 facetime 功能
class MobilePhone:
    screen = 'Touch'
    camera = 'Megapixel'

    def __init__(self, brand, color, model='OrdinaryPhone', system_type='Android', cellphone_number=None, caller=None,
                 answer=None):
        self.brand = brand
        self.color = color
        self.model = model
        self.system_type = system_type
        self.cellphone_number = cellphone_number
        self.caller = caller
        self.answer = answer
        print('品牌：{}，颜色：{}，类型：{}，系统：{}'.format(brand, color, model, system_type))

    def card(self):
        return self.cellphone_number

    def call(self):
        print('{}使用{}号码打电话给{}'.format(self.caller, self.card(), self.answer))

    @staticmethod
    def gap():
        print('\\\\', '*' * 69)


class SmartPhone(MobilePhone):
    def __init__(self, brand, color, record_content, model='OrdinaryPhone', system_type='Android',
                 cellphone_number=None, caller=None,
                 answer=None):
        super(SmartPhone, self).__init__(brand, color, model='OrdinaryPhone', system_type='Android')
        self.brand = brand
        self.color = color
        self.record_content = record_content
        self.model = model
        self.system_type = system_type
        self.cellphone_number = cellphone_number
        self.caller = caller
        self.answer = answer

    def call(self, record=False):
        super(SmartPhone, self).call()
        if record:
            print('{}使用{}号码打电话给{}说{}'.format(self.caller, self.card(), self.answer, self.record()))

    def record(self):
        return self.record_content


class ApplePhone(SmartPhone):
    def call(self, record=False, facetime=False):
        super(ApplePhone, self).call(record)
        if facetime:
            print('{}使用{}号码打电话给{}{}'.format(self.caller, self.card(), self.answer, self.facetime()))

    def facetime(self):
        return '正在FaceTime通话'


my_phone1 = MobilePhone('树树', '红色', cellphone_number=13888888888, caller='刀刀', answer='树树')
my_phone1.gap()
my_phone2 = SmartPhone('树树', '红色', '我想你', cellphone_number=13888888888, caller='刀刀', answer='树树')
my_phone2.call(record=True)
my_phone1.gap()
my_phone3 = ApplePhone('树树', '红色', '我想你', cellphone_number=13888888888, caller='刀刀', answer='树树')
my_phone3.call(record=False, facetime=True)
my_phone1.gap()
my_phone3.call(record=True, facetime=False)
my_phone1.gap()
my_phone3.call(record=True, facetime=True)
my_phone1.gap()
my_phone3.call(record=False, facetime=False)


# 3、定义一个 ExcelManual 类。具有获取 sheet 表单， 读取单元格 和 修改单元格功能。
class ExcelManual:
    def __init__(self, path):
        self.path = path

    def open_excle(self):
        """
        加载Excle文件
        :return: 返回Excle文件这个对象
        """
        load_excle = load_workbook(self.path)
        return load_excle

    def create_sheet(self):
        """
        创建表单
        :return: 返回已创建表单的新Excle文件
        """
        file_excle = self.open_excle()
        new_sheet = file_excle.create_sheet('Shushu')
        print(new_sheet)  # 创建表单
        # file_excle.save(self.path) #避免下方调用创建太多表单，暂时不保存创建后的Excle文件
        return file_excle

    def get_sheet(self):
        """
        获取表单
        :return: self.open_excle()['Sheet1']指定名称的表单
        """
        print(self.create_sheet().worksheets)  # 获取所有表单
        print(self.create_sheet().sheetnames)  # 获取所有表单名字
        print(self.create_sheet().worksheets[0])  # 获取指定位置(索引)的表单
        print(self.create_sheet()['Sheet1'])  # 获取指定名字的表单
        print(self.create_sheet().active)  # 获取激活的表单
        return self.create_sheet()['Sheet1']

    def get_cell(self):
        """
        读取单元格
        :return: None
        """
        print(self.get_sheet().cell(1, 1).value)  # 使用cell方法获取单元格的值
        print(self.get_sheet()['A2'].value)  # 使用行列值获取对应单元格的值

    def get_max_row_col(self):
        print(self.get_sheet().max_row)  # 获取表单的最大行数
        print(self.get_sheet().max_column)  # 获取表单的最大列数
        self.create_sheet().close()

    def write_sheet(self):
        """
        创建Excle文件并修改指定单元格的值和保存文件
        :return:None
        """
        create_excle = Workbook()
        active_sheet = create_excle.active
        active_sheet['A1'] = 'Hello'
        active_sheet.cell(1, 2, value='Daodao')
        active_sheet.cell(1, 3).value = 'Shushu'
        create_excle.save('demo1.xlsx')
        create_excle.close()
        return create_excle, active_sheet

    # TODO copy_sheet还没有理解是怎么使用的，需要继续跟进
    def copy_sheet(self):
        create_excle = self.write_sheet()[0]
        active_sheet = self.write_sheet()[1]
        create_excle.copy_worksheet(active_sheet)
        create_excle.save('demo1.xlsx')


my_excle = ExcelManual('demo.xlsx')
# my_excle.create_sheet()
my_excle.get_max_row_col()
# my_excle.write_sheet()
