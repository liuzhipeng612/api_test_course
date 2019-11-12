#!/usr/bin/env python
# coding=UTF-8
"""
@Author: Cautious Eureka
@Contact: Eureka@Faisen.com
@Project: cautious_eureka
@File: cw0618.py
@Time: 2019-06-24 23:48
@Desc: Python full stack automation test engineer(Cautious Eureka)
"""
from openpyxl import load_workbook, Workbook


# 一、使用 openpyxl 完善上节课的 ExcelManal  操作类。
class ExcelManal:
    def __init__(self, path):
        self.path = path
        self.load_excle = None
        pass

    def open_excle(self):
        self.load_excle = load_workbook(self.path)
        return self.load_excle

    # 1、选择表单功能
    def choose_sheet(self):
        self.open_excle()
        view_sheet = self.load_excle.worksheets
        sheet_names = self.load_excle.sheetnames
        active_sheet = self.load_excle.active
        select_sheet = self.load_excle.worksheets[0]
        assign_sheet = self.load_excle['Sheet']
        return view_sheet, sheet_names, active_sheet, select_sheet, assign_sheet

    # 2、读取一个单元格的数据功能
    def read_cell(self):
        get_sheet = self.choose_sheet()[2]
        cell_value = get_sheet['A1'].value
        cell_value = get_sheet.cell(1, 1).value
        return cell_value

    # 3、读取一行数据的功能
    def read_row(self):
        get_sheet = self.choose_sheet()[2]
        row_data = []
        for i in get_sheet[1]:
            row_data += i.coordinate, i.value
        return row_data

    # 4、读取表单中所有数据功能
    def read_sheet(self):
        get_sheet = self.choose_sheet()[2]
        for j in get_sheet.iter_rows():
            for cell in j:
                print(cell.coordinate, cell.value)

    # 5、往单元格中写入数据功能和保存数据
    def write_cell(self):
        get_sheet = self.choose_sheet()[2]
        get_sheet['A1'] = 'Shushu'
        get_sheet.cell(1, 1).value = 'Shushu'
        self.load_excle.save(self.path)
        self.load_excle.close()


my_excle = ExcelManal('demo.xlsx')
# my_excle.read_row()
# my_excle.read_sheet()
my_excle.write_cell()


# 二、

class one_excle:
    def __init__(self, path):
        self.path = path

    def create_excle(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['url1', 'data1', '请求方法1', '预期结果1', '等数据1'])
        ws.append(['url2', 'data2', '请求方法2', '预期结果2', '等数据2'])
        ws.append(['url3', 'data3', '请求方法3', '预期结果3', '等数据3'])
        ws.append(['url4', 'data4', '请求方法4', '预期结果4', '等数据4'])
        ws.append(['url5', 'data5', '请求方法5', '预期结果5', '等数据5'])
        wb.save(self.path)
        wb.close()
        wb1 = load_workbook(self.path)
        ws1 = wb1.active
        for j in ws1.iter_rows():
            for cell in j:
                print(cell.coordinate, cell.value)
        ws1['F1'] = 1
        ws1['F2'] = 2
        ws1['F3'] = 3
        ws1['F4'] = 4
        ws1['F5'] = 5
        wb1.save(self.path)
        wb1.close()


# 2、保存测试用例数据 5 条，包含 url, data, 请求方法，预期结果 等数据。
# 3、 读取 刚才的 excel 文件，分别打印每一行的数据
# 4、创建一个实际结果列，分别填入对应的实际结果。
my_one = one_excle('demo1.xlsx')
my_one.create_excle()
