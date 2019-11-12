#!/usr/bin/env python
# coding=UTF-8
"""
@Author: Cautious Eureka
@Contact: Eureka@Faisen.com
@Project: cautious_eureka
@File: test_cw0622_method.py
@Time: 2019-06-28 01:08
@Desc: Python full stack automation test engineer(Cautious Eureka)
"""
import unittest

from openpyxl import load_workbook


class LoadOneExcel:
    def __init__(self, path='demo.xlsx'):
        self.path = path

    def open_excel(self):
        wb = load_workbook(self.path)
        ws = wb.active
        return ws

    def get_value1(self):
        return self.open_excel().cell(2, 3, ).value

    def get_value2(self):
        return self.open_excel().cell(1, 2).value


class TestExcel(unittest.TestCase, LoadOneExcel):
    def __init__(self, methodname):
        super().__init__(methodName=methodname)
        self.path = 'demo.xlsx'

    def test_none(self):
        self.assertIsNone(self.get_value1())

    def test_equal(self):
        self.assertMultiLineEqual(self.get_value2(), '长沙')


if __name__ == '__main__':
    unittest.main()
