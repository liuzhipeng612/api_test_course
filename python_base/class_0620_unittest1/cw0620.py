#!/usr/bin/env python
# coding=UTF-8
"""
@Author: Cautious Eureka
@Contact: Eureka@Faisen.com
@Project: cautious_eureka
@File: cw0620test.py
@Time: 2019-06-26 00:31
@Desc: Python full stack automation test engineer(Cautious Eureka)
"""
from openpyxl import load_workbook, Workbook
import unittest


# 1、封装一个方法或者类，读取 excel 文件里面的数据
class OneExcel:
    def __init__(self, path):
        self.path = path

    def open_excel(self):
        wb = Workbook()
        ws = wb.active
        ws.append(range(10))
        wd = tuple(ws.iter_rows(values_only=True))
        wd1 = ''
        for i in wd[0]:
            wd1 += str(eval('i'))
        wb.save(self.path)
        print(wd1)


my_excel = OneExcel('demo.xlsx')
my_excel.open_excel()


# 2、测试用例1：断言文件 第2行 第三列是否为空
class TestExcel(unittest.TestCase):
    def __init__(self, methodname):
        super().__init__(methodName=methodname)
        self.path = 'demo.xlsx'

    def setUp(self):
        wb = load_workbook(self.path)
        ws = wb.active
        wd = ws.cell(2, 3).value
        wd1 = ws.cell(1, 2).value
        return wd, wd1

    def test_none(self):
        self.assertIsNone(self.setUp()[0])

    # 3、测试用例2：断言文件 第1行 第二列是否和长沙相等。

    def test_equal(self):
        self.assertMultiLineEqual(self.setUp()[1], '长沙')


if __name__ == '__main__':
    unittest.main()
